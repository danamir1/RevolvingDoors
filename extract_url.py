from openpyxl import load_workbook, Workbook
import pandas as pd

sadna_clients = 'clients_lobbyists.csv'
sadna_lobbyists = 'data_lobbyists.csv'
ours = 'Lobbyist.xlsx'


# add url
def extract_url(file_name=ours):
    """
    Extract the url of each lobbyist and saves the updated xlsx
    :param file_name: path to xlsx
    """
    wb = load_workbook(file_name)
    ws = wb.get_sheet_by_name('Sheet1')
    column = 1
    target = 5
    start_row = 2
    end_row = 161
    ws.cell(row=1, column=target).value = 'link'
    ws.cell(row=1, column=target + 1).value = 'LobbyistID'
    for row in range(start_row, end_row + 1):
        url = ws.cell(row=row, column=column).hyperlink.target
        # print(url)
        ws.cell(row=row, column=target).value = url

        id = int(url.split('=')[-1])
        ws.cell(row=row, column=target + 1).value = id
    wb.save(file_name)


def validate_ids(sadna=sadna_clients, ours=ours):
    sadna_df = pd.read_csv(sadna)  # , sheet_name='Sheet1')
    ours_df = pd.read_excel(ours, sheet_name='Sheet1')
    sadna_ids = set(sadna_df['LobbyistID'])
    ours_ids = set(ours_df['LobbyistID'])

    print('same id\'s:', sadna_ids == ours_ids)
    print('ours ids:', len(ours_ids), 'sadna ids:', len(sadna_ids))
    print('we have and sadna doesnt:', len(ours_ids - sadna_ids))
    print('sadna has and we dont:', len(sadna_ids - ours_ids))
    total = sadna_ids.union(ours_ids)
    dif = set(sadna_ids - ours_ids).union(ours_ids - sadna_ids)
    print('total ids:', len(total), 'total mismatches:', len(dif))


def organize_pairs(sadna_clients_path=sadna_clients, sadna_lobbyists_path=sadna_lobbyists,
                   ours_path=ours):
    sadna_clients = pd.read_csv(sadna_clients_path)
    sadna_lobbyists = pd.read_csv(sadna_lobbyists_path)
    ours = pd.read_excel(ours_path, sheet_name='Sheet1')
    df = pd.DataFrame(columns=['company id', 'company name', 'lobbyist id', 'lobbyist name',
                               'exists online', 'representation', 'lobbyist corporation name',
                               'lobbyist corporation id', 'lobbyist party', 'lobbyist link'])
    # print(sadna)
    for index, row in sadna_clients.iterrows():
        lobbyist_id = row['LobbyistID']
        lobbyist = ours[ours.LobbyistID == lobbyist_id]
        # print(lobbyist['שם השדלן'], not lobbyist.empty)
        rep = row['ClientsNames'].split('- ')[-1]
        if lobbyist.empty:
            lobbyist = sadna_lobbyists[sadna_lobbyists.LobbyistID == lobbyist_id].iloc[0]
            party = lobbyist['IsMemberInFaction']
            if party != party:
                # value is nan
                party = 'לא'
            df.loc[index] = [row['ClientID'], row['Name'], lobbyist_id, lobbyist['FullName'], False,
                             'NA', lobbyist['CorporationName'], lobbyist['CorpNumber'],
                             party, 'NA']

        else:
            lobbyist = lobbyist.iloc[0]
            df.loc[index] = [row['ClientID'], row['Name'], lobbyist_id, lobbyist['שם השדלן'],
                             True, rep, lobbyist['שם התאגיד'], lobbyist['מספר התאגיד'],
                             lobbyist['חברות בגוף בוחר'], lobbyist['link']]

    df.to_excel('lobbyist-company.xlsx')
    df.to_csv('lobbyist-company.csv')


if __name__ == '__main__':
    # extract_url()
    # validate_ids()
    organize_pairs()
